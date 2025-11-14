"""Excel export utilities."""
import os
import time
import pandas as pd
from file_manager import FileManager
from config import DEPARTMENTS, TARGET_SEMESTERS, PRE_MID, POST_MID
from excel_loader import ExcelLoader
from openpyxl.styles import PatternFill

class ExcelExporter:
    """Handles exporting of timetables to Excel files."""
    
    def __init__(self, data_frames, schedule_generator):
        self.dfs = data_frames
        self.schedule_gen = schedule_generator
        # Expanded vibrant pastel palette (readable on black text)
        self._palette = [
            "FFCDD2","F8BBD0","E1BEE7","D1C4E9","C5CAE9","BBDEFB","B3E5FC","B2EBF2",
            "B2DFDB","C8E6C9","DCEDC8","F0F4C3","FFF9C4","FFECB3","FFE0B2","FFCCBC",
            "D7CCC8","CFD8DC",
            "F28B82","F7A1C4","B39DDB","9FA8DA","90CAF9","81D4FA","80DEEA","80CBC4",
            "A5D6A7","C5E1A5","E6EE9C","FFF59D","FFE082","FFCC80","FFAB91",
            "AED581","81C784","4DD0E1","4FC3F7","9575CD","F48FB1"
        ]
        # Deterministic color mapping per exported workbook
        self._course_color_map = {}
    
    def _course_from_cell(self, val: str) -> str:
        """Extract a course identifier from a cell value."""
        if val is None:
            return ""
        s = str(val).strip()
        if not s or s == "-" or s.upper() == "FREE" or s.upper().startswith("LUNCH"):
            return ""
        # Common patterns: "CS161", "CS161 (Lab)", "CS161-Lab", "CS161: L"
        # Take up to first space or '(' or ':' or '-'
        for sep in [" (", " -", ":", " "]:
            if sep in s:
                s = s.split(sep)[0]
                break
        return s.strip()
    
    def _color_for_course(self, course: str) -> str:
        """Pick a stable color for the course within the current export."""
        if not course:
            return None
        if course not in self._course_color_map:
            idx = len(self._course_color_map) % len(self._palette)
            self._course_color_map[course] = self._palette[idx]
        return self._course_color_map[course]
    
    def _apply_color_coding(self, worksheet, schedule_df, start_row=1, start_col=1):
        """Apply background colors to timetable cells based on course code."""
        # Build mapping using shared cache to keep colors consistent across sheets
        course_to_color = {}
        for day in schedule_df.index:
            for slot in schedule_df.columns:
                val = schedule_df.loc[day, slot]
                course = self._course_from_cell(val)
                if course and course not in course_to_color:
                    course_to_color[course] = self._color_for_course(course)
        # Apply fills
        # Dataframe written starting at (row=start_row, col=start_col), with header row and index col
        header_rows = 1
        index_cols = 1
        nrows = len(schedule_df.index)
        ncols = len(schedule_df.columns)
        for r in range(nrows):
            for c in range(ncols):
                cell = worksheet.cell(row=start_row + header_rows + r, column=start_col + index_cols + c)
                val = cell.value
                course = self._course_from_cell(val)
                if course and course in course_to_color:
                    color = course_to_color[course]
                    try:
                        cell.fill = PatternFill(fill_type="solid", fgColor=color)
                    except Exception:
                        pass
    
    def _get_course_details_for_session(self, semester, department, session_type):
        """Get course details for a specific department and session.
        Validates that expected courses from division logic match what should be scheduled."""
        try:
            # Get all semester courses
            sem_courses_all = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses_all.empty:
                return pd.DataFrame()
            
            # Parse LTPSC
            sem_courses_parsed = ExcelLoader.parse_ltpsc(sem_courses_all)
            if sem_courses_parsed.empty:
                return pd.DataFrame()
            
            # Filter for department
            if 'Department' in sem_courses_parsed.columns:
                dept_mask = sem_courses_parsed['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                dept_courses = sem_courses_parsed[dept_mask].copy()
            else:
                dept_courses = sem_courses_parsed.copy()
            
            if dept_courses.empty:
                return pd.DataFrame()
            
            # Divide by session
            pre_mid_courses, post_mid_courses = ExcelLoader.divide_courses_by_session(dept_courses, department, all_sem_courses=sem_courses_parsed)
            
            # Select appropriate session
            if session_type == PRE_MID:
                session_courses = pre_mid_courses
            else:
                session_courses = post_mid_courses
            
            if session_courses.empty:
                print(f"    WARNING: No courses assigned to {department} {session_type} session")
                return pd.DataFrame()
            
            # Prepare summary data
            summary_columns = ['Course Code', 'Course Name', 'Instructor', 'LTPSC', 'Lectures_Per_Week', 'Tutorials_Per_Week', 'Labs_Per_Week', 'Room Allocated', 'Lab Room Allocated', 'Combined Class']
            available_cols = [col for col in summary_columns if col in session_courses.columns]
            
            summary_df = session_courses[available_cols].copy()
            # Ensure Combined Class column exists
            if 'Combined Class' not in summary_df.columns:
                summary_df['Combined Class'] = 'NO'
            # Ensure Room Allocated columns exist
            if 'Room Allocated' not in summary_df.columns:
                summary_df['Room Allocated'] = ''
            if 'Lab Room Allocated' not in summary_df.columns:
                summary_df['Lab Room Allocated'] = ''
            
            # Format counts as "allocated/required" (e.g., "2/3" means 2 allocated out of 3 required)
            if 'Course Code' in summary_df.columns:
                for idx, row in summary_df.iterrows():
                    course_code = str(row.get('Course Code', '')).strip()
                    if course_code and course_code != 'nan' and course_code:
                        # Get required (expected) counts from course data
                        # Use 0 as default if column doesn't exist or value is missing
                        required_lectures = 0
                        required_tutorials = 0
                        required_labs = 0
                        
                        if 'Lectures_Per_Week' in summary_df.columns:
                            required_lectures = pd.to_numeric(row.get('Lectures_Per_Week', 0), errors='coerce')
                            if pd.isna(required_lectures):
                                required_lectures = 0
                            required_lectures = int(required_lectures)
                        
                        if 'Tutorials_Per_Week' in summary_df.columns:
                            required_tutorials = pd.to_numeric(row.get('Tutorials_Per_Week', 0), errors='coerce')
                            if pd.isna(required_tutorials):
                                required_tutorials = 0
                            required_tutorials = int(required_tutorials)
                        
                        if 'Labs_Per_Week' in summary_df.columns:
                            required_labs = pd.to_numeric(row.get('Labs_Per_Week', 0), errors='coerce')
                            if pd.isna(required_labs):
                                required_labs = 0
                            required_labs = int(required_labs)
                        
                        # Get actual allocated counts from schedule generator
                        actual = self.schedule_gen.get_actual_allocations(semester, department, session_type, course_code)
                        actual_lectures = actual.get('lectures', 0)
                        actual_tutorials = actual.get('tutorials', 0)
                        actual_labs = actual.get('labs', 0)
                        
                        # Combined classes disabled; always NO
                        combined_used = False
                        # Room from schedule_gen allocations
                        room_alloc = actual.get('room', '')
                        lab_room_alloc = actual.get('lab_room', '')
                        # Format as "allocated/required"
                        if 'Lectures_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Lectures_Per_Week'] = f"{actual_lectures}/{required_lectures}"
                        if 'Tutorials_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Tutorials_Per_Week'] = f"{actual_tutorials}/{required_tutorials}"
                        if 'Labs_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Labs_Per_Week'] = f"{actual_labs}/{required_labs}"
                        # Mark combined usage
                        summary_df.at[idx, 'Combined Class'] = 'YES' if combined_used else 'NO'
                        if 'Room Allocated' in summary_df.columns:
                            summary_df.at[idx, 'Room Allocated'] = room_alloc
                        if 'Lab Room Allocated' in summary_df.columns:
                            summary_df.at[idx, 'Lab Room Allocated'] = lab_room_alloc
            
            # Validate: Check if any courses have zero LTPSC (should still be included but may not schedule)
            # Note: Now checking actual allocated values from the formatted strings
            if 'Course Code' in summary_df.columns and 'Lectures_Per_Week' in summary_df.columns:
                # Extract actual values from "allocated/required" format for validation
                actual_lectures_list = []
                actual_tutorials_list = []
                actual_labs_list = []
                
                for idx, row in summary_df.iterrows():
                    # Extract allocated value from "allocated/required" format
                    lec_str = str(row.get('Lectures_Per_Week', '0/0'))
                    tut_str = str(row.get('Tutorials_Per_Week', '0/0'))
                    lab_str = str(row.get('Labs_Per_Week', '0/0'))
                    
                    # Parse "allocated/required" format
                    try:
                        actual_lec = int(lec_str.split('/')[0]) if '/' in lec_str else 0
                        actual_tut = int(tut_str.split('/')[0]) if '/' in tut_str else 0
                        actual_lab = int(lab_str.split('/')[0]) if '/' in lab_str else 0
                    except:
                        actual_lec = 0
                        actual_tut = 0
                        actual_lab = 0
                    
                    actual_lectures_list.append(actual_lec)
                    actual_tutorials_list.append(actual_tut)
                    actual_labs_list.append(actual_lab)
                
                # Check for zero LTPSC
                zero_ltpsc_mask = (
                    (pd.Series(actual_lectures_list) == 0) &
                    (pd.Series(actual_tutorials_list) == 0) &
                    (pd.Series(actual_labs_list) == 0)
                )
                zero_ltpsc = summary_df[zero_ltpsc_mask]
                if not zero_ltpsc.empty:
                    zero_codes = zero_ltpsc['Course Code'].dropna().tolist()
                    print(f"    INFO: {len(zero_codes)} courses with 0-0-0 LTPSC in {department} {session_type}: {', '.join(zero_codes)}")
            
            # Rename columns for better display
            column_rename = {
                'Lectures_Per_Week': 'Lectures/Week',
                'Tutorials_Per_Week': 'Tutorials/Week',
                'Labs_Per_Week': 'Labs/Week',
                'Instructor': 'Faculty'
            }
            summary_df = summary_df.rename(columns=column_rename)
            
            return summary_df
            
        except Exception as e:
            print(f"    WARNING: Could not generate course details: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def export_semester_timetable(self, semester):
        """Export timetable for a specific semester."""
        print(f"\n{'='*60}")
        print(f"GENERATING SEMESTER {semester} TIMETABLE")
        print(f"{'='*60}")
        # Reset color map for each workbook to keep palette consistent within file
        self._course_color_map = {}
        
        filename = f"sem{semester}_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        
        # Attempt to open writer, handle PermissionError (file locked by Excel)
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError as pe:
            print(f"\nWARNING: Cannot write to {filepath} (Permission denied / file may be open).")
            timestamp = int(time.time())
            alt_filename = f"sem{semester}_timetable_{timestamp}.xlsx"
            alt_filepath = FileManager.get_output_path(alt_filename)
            print(f"Attempting alternative filename: {alt_filename}")
            try:
                writer = pd.ExcelWriter(alt_filepath, engine='openpyxl')
                filepath = alt_filepath
                filename = alt_filename
            except Exception as e:
                print(f"\nFAILED: Could not create {filename}: {e}")
                import traceback
                traceback.print_exc()
                return False
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        # Use the opened writer (writer variable guaranteed) in a context manager
        try:
            with writer as w:
                print(f"Creating {filename}...")
                
                # Write Course_Summary first so workbook always has at least one visible sheet
                try:
                    self._add_course_summary(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not write initial Course_Summary: {e}")
                
                # Generate schedules for each department and session
                department_count = 0
                for department in DEPARTMENTS:
                    print(f"\nProcessing {department}:")
                    
                    # Pre-Mid session
                    print(f"  {PRE_MID} session...")
                    try:
                        pre_mid_schedule = self.schedule_gen.generate_department_schedule(semester, department, PRE_MID)
                    except Exception as e:
                        print(f"    ERROR generating {department} {PRE_MID}: {e}")
                        pre_mid_schedule = self.schedule_gen._initialize_schedule()
                    
                    if pre_mid_schedule is not None:
                        sheet_name = f"{department}_{PRE_MID}"
                        clean_schedule = pre_mid_schedule.replace('Free', '-')
                        
                        # Write schedule first
                        clean_schedule.to_excel(w, sheet_name=sheet_name, index=True, startrow=0)
                        
                        # Apply color coding to schedule grid
                        try:
                            ws = w.sheets[sheet_name]
                            self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                        except Exception as e:
                            print(f"    WARNING: Could not apply color coding to {sheet_name}: {e}")
                        
                        # Get course details for this session
                        course_details = self._get_course_details_for_session(semester, department, PRE_MID)
                        
                        # Add course details below the schedule
                        if not course_details.empty:
                            # Calculate starting row (schedule rows + header + 2 blank rows)
                            start_row = len(clean_schedule) + 3
                            
                            # Write a header for course details section
                            worksheet = w.sheets[sheet_name]
                            worksheet.cell(row=start_row, column=1, value="COURSE DETAILS:")
                            
                            # Write course details table
                            course_details.to_excel(w, sheet_name=sheet_name, index=False, startrow=start_row+1)
                        
                        print(f"    SUCCESS: {sheet_name} created with course details")
                        department_count += 1
                    else:
                        print(f"    FAILED: {department}_{PRE_MID}")
                    
                    # Post-Mid session  
                    print(f"  {POST_MID} session...")
                    try:
                        post_mid_schedule = self.schedule_gen.generate_department_schedule(semester, department, POST_MID)
                    except Exception as e:
                        print(f"    ERROR generating {department} {POST_MID}: {e}")
                        post_mid_schedule = self.schedule_gen._initialize_schedule()
                    
                    if post_mid_schedule is not None:
                        sheet_name = f"{department}_{POST_MID}"
                        clean_schedule = post_mid_schedule.replace('Free', '-')
                        
                        # Write schedule first
                        clean_schedule.to_excel(w, sheet_name=sheet_name, index=True, startrow=0)
                        
                        # Apply color coding to schedule grid
                        try:
                            ws = w.sheets[sheet_name]
                            self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                        except Exception as e:
                            print(f"    WARNING: Could not apply color coding to {sheet_name}: {e}")
                        
                        # Get course details for this session
                        course_details = self._get_course_details_for_session(semester, department, POST_MID)
                        
                        # Add course details below the schedule
                        if not course_details.empty:
                            # Calculate starting row (schedule rows + header + 2 blank rows)
                            start_row = len(clean_schedule) + 3
                            
                            # Write a header for course details section
                            worksheet = w.sheets[sheet_name]
                            worksheet.cell(row=start_row, column=1, value="COURSE DETAILS:")
                            
                            # Write course details table
                            course_details.to_excel(w, sheet_name=sheet_name, index=False, startrow=start_row+1)
                        
                        print(f"    SUCCESS: {sheet_name} created with course details")
                        department_count += 1
                    else:
                        print(f"    FAILED: {department}_{POST_MID}")
                
                print(f"\nSUCCESS: Created {filename}")
                print(f"  - {department_count} department schedules")
                print(f"  - Course summary sheet")
            
            return True
            
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _add_course_summary(self, writer, semester):
        """Add course information summary. Always create the Course_Summary sheet (may be empty).
        Adds LTPSC validity check for all courses."""
        try:
            # Prepare empty default summary (columns if available)
            default_cols = ['Course Code', 'Course Name', 'LTPSC', 'Credits']
            summary_df = pd.DataFrame(columns=default_cols)

            ltpsc_valid_col = []
            all_valid = True

            if 'course' in self.dfs:
                course_df = self.dfs['course']
                if 'Semester' in course_df.columns:
                    temp_df = course_df.copy()
                    temp_df['Semester'] = pd.to_numeric(temp_df['Semester'], errors='coerce')
                    sem_courses = temp_df[temp_df['Semester'] == semester]

                    if not sem_courses.empty:
                        available_cols = [col for col in default_cols if col in sem_courses.columns]
                        summary_df = sem_courses[available_cols].copy()
                        # Check LTPSC validity for each course
                        for idx, row in summary_df.iterrows():
                            ltpsc_val = str(row.get('LTPSC', '')).strip()
                            valid = False
                            if ltpsc_val and '-' in ltpsc_val:
                                parts = ltpsc_val.split('-')
                                if len(parts) >= 3:
                                    try:
                                        float(parts[0])
                                        float(parts[1])
                                        float(parts[2])
                                        valid = True
                                    except Exception:
                                        valid = False
                            ltpsc_valid_col.append(valid)
                            if not valid:
                                all_valid = False
                        summary_df['LTPSC_Valid'] = ltpsc_valid_col
                        print(f"SUCCESS: Added Course_Summary sheet with {len(summary_df)} courses")
                    else:
                        print(f"WARNING: No courses found for semester {semester}; writing empty Course_Summary")
                else:
                    print("WARNING: 'Semester' column not found in course data; writing empty Course_Summary")
            else:
                print("WARNING: 'course' data frame not found; writing empty Course_Summary")

            # Add a message row at the top
            from pandas import DataFrame
            msg = "All courses follow LTPSC structure." if all_valid and not summary_df.empty else "Some courses do NOT follow LTPSC structure."
            msg_df = DataFrame({'Course Code': [msg]})
            # Write message row, then summary directly to the existing writer
            msg_df.to_excel(writer, sheet_name='Course_Summary', index=False, header=False, startrow=0)
            summary_df.to_excel(writer, sheet_name='Course_Summary', index=False, startrow=2)
        except Exception as e:
            print(f"FAILED: Could not add course summary: {e}")
    
    def export_semester7_timetable(self):
        """Export special unified timetable for 7th semester with baskets.
        Creates:
        1. Main timetable showing baskets (7B1, 7B2, 7B3, 7B4) - 9:00 AM to 5:30 PM only, 2 classes per basket
        2. Basket assignments sheet (which courses go to which baskets)"""
        semester = 7
        print(f"\n{'='*60}")
        print(f"GENERATING SEMESTER {semester} UNIFIED TIMETABLE (BASKETS)")
        print(f"{'='*60}")
        # Reset color map for each workbook
        self._course_color_map = {}
        
        filename = f"sem{semester}_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        
        # Attempt to open writer
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError as pe:
            print(f"\nWARNING: Cannot write to {filepath} (Permission denied / file may be open).")
            import time
            timestamp = int(time.time())
            alt_filename = f"sem{semester}_timetable_{timestamp}.xlsx"
            alt_filepath = FileManager.get_output_path(alt_filename)
            print(f"Attempting alternative filename: {alt_filename}")
            try:
                writer = pd.ExcelWriter(alt_filepath, engine='openpyxl')
                filepath = alt_filepath
                filename = alt_filename
            except Exception as e:
                print(f"\nFAILED: Could not create {filename}: {e}")
                import traceback
                traceback.print_exc()
                return False
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        try:
            with writer as w:
                print(f"Creating {filename}...")
                
                # Get 7th semester courses
                if 'course' not in self.dfs:
                    print("ERROR: Course data not found")
                    return False
                
                course_df = self.dfs['course']
                if 'Semester' not in course_df.columns:
                    print("ERROR: Semester column not found")
                    return False
                
                # Filter 7th semester courses
                temp_df = course_df.copy()
                temp_df['Semester'] = pd.to_numeric(temp_df['Semester'], errors='coerce')
                sem7_courses = temp_df[temp_df['Semester'] == semester].copy()
                
                # Separate baskets and non-basket courses
                if 'Course Code' not in sem7_courses.columns:
                    print("ERROR: Course Code column not found")
                    return False
                
                # Identify baskets (pattern: 7B1, 7B2, 7B3, 7B4, etc.)
                basket_mask = sem7_courses['Course Code'].astype(str).str.match(r'^7B\d+', na=False)
                baskets = sem7_courses[basket_mask].copy()
                non_basket_courses = sem7_courses[~basket_mask].copy()
                
                print(f"Found {len(baskets)} baskets: {', '.join(baskets['Course Code'].astype(str).tolist()) if not baskets.empty else 'None'}")
                print(f"Found {len(non_basket_courses)} non-basket courses")
                
                # 1. Generate unified timetable with baskets
                # For 7th semester: classes only from 9:00 AM to 5:30 PM
                from config import DAYS, TEACHING_SLOTS, LECTURE_DURATION
                from config import LUNCH_SLOTS
                
                # Filter slots to only include 9:00 AM to 5:30 PM (17:30)
                # Slots start from '09:00-09:30' and go until '17:00-17:30'
                sem7_slots = [s for s in TEACHING_SLOTS if s >= '09:00-09:30' and s <= '17:00-17:30']
                
                schedule = pd.DataFrame(index=DAYS, columns=sem7_slots)
                for day in DAYS:
                    for slot in sem7_slots:
                        schedule.loc[day, slot] = 'Free'
                
                # Mark lunch slots
                for day in DAYS:
                    for lunch_slot in LUNCH_SLOTS:
                        if lunch_slot in schedule.columns:
                            schedule.loc[day, lunch_slot] = 'LUNCH BREAK'
                
                # Schedule baskets - assign each basket to different time slots
                # Each basket gets 2 lectures per week (for 7th semester)
                import random
                basket_codes = baskets['Course Code'].astype(str).tolist() if not baskets.empty else []
                
                # Schedule each basket with 2 lectures per week
                for basket_code in basket_codes:
                    scheduled = 0
                    attempts = 0
                    max_attempts = 100
                    
                    while scheduled < 2 and attempts < max_attempts:
                        attempts += 1
                        day = random.choice(DAYS)
                        # Avoid lunch slots - use only sem7_slots
                        available_slots = [s for s in sem7_slots if s not in LUNCH_SLOTS]
                        if not available_slots:
                            continue
                        
                        start_slot = random.choice(available_slots)
                        try:
                            start_idx = sem7_slots.index(start_slot)
                            end_idx = start_idx + LECTURE_DURATION
                            if end_idx > len(sem7_slots):
                                continue
                            slots = sem7_slots[start_idx:end_idx]
                            
                            # Check if all slots are free
                            if all(schedule.loc[day, s] == 'Free' for s in slots):
                                # Check if any slot is lunch
                                if any(s in LUNCH_SLOTS for s in slots):
                                    continue
                                
                                # Assign basket to these slots
                                for slot in slots:
                                    schedule.loc[day, slot] = basket_code
                                scheduled += 1
                        except (ValueError, IndexError):
                            continue
                
                # Write main timetable
                clean_schedule = schedule.replace('Free', '-')
                clean_schedule.to_excel(w, sheet_name='Timetable', index=True, startrow=0)
                
                # Apply color coding
                try:
                    ws = w.sheets['Timetable']
                    self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                except Exception as e:
                    print(f"    WARNING: Could not apply color coding: {e}")
                
                print(f"    SUCCESS: Main timetable created with {len(basket_codes)} baskets")
                
                # 2. Create basket assignments sheet
                # Check if there's a "7th sem " sheet with basket assignments
                basket_assignments = pd.DataFrame(columns=['Basket Code', 'Course Code', 'Course Name', 'Department', 'LTPSC', 'Credits', 'Instructor'])
                
                # Look for 7th sem sheet in data_frames
                sem7_sheet_key = None
                for key in self.dfs.keys():
                    key_lower = key.lower()
                    # Match patterns like "course_7th_sem", "7th_sem", etc.
                    if ('7th' in key_lower and 'sem' in key_lower) or key_lower == '7th_sem_':
                        sem7_sheet_key = key
                        break
                
                if sem7_sheet_key and sem7_sheet_key in self.dfs:
                    sem7_sheet_df = self.dfs[sem7_sheet_key]
                    print(f"    Found 7th semester sheet: {sem7_sheet_key} with {len(sem7_sheet_df)} courses")
                    
                    # Map columns from the sheet to our format
                    basket_col = None
                    course_code_col = None
                    course_name_col = None
                    faculty_col = None
                    
                    for col in sem7_sheet_df.columns:
                        col_lower = str(col).lower()
                        if 'basket' in col_lower:
                            basket_col = col
                        elif 'course code' in col_lower:
                            course_code_col = col
                        elif col_lower == 'course' or 'course name' in col_lower:
                            course_name_col = col
                        elif 'faculty' in col_lower or 'instructor' in col_lower:
                            faculty_col = col
                    
                    if basket_col and course_code_col:
                        # Build basket assignments dataframe
                        for _, row in sem7_sheet_df.iterrows():
                            basket_code = str(row.get(basket_col, '')).strip()
                            course_code = str(row.get(course_code_col, '')).strip()
                            course_name = str(row.get(course_name_col, '')).strip() if course_name_col else ''
                            instructor = str(row.get(faculty_col, '')).strip() if faculty_col else ''
                            
                            # Try to get additional info from main course data if available
                            dept = ''
                            ltpsc = ''
                            credits = ''
                            
                            if not course_df.empty and 'Course Code' in course_df.columns:
                                course_match = course_df[course_df['Course Code'].astype(str) == course_code]
                                if not course_match.empty:
                                    match_row = course_match.iloc[0]
                                    dept = str(match_row.get('Department', '')) if 'Department' in match_row else ''
                                    ltpsc = str(match_row.get('LTPSC', '')) if 'LTPSC' in match_row else ''
                                    credits = str(match_row.get('Credits', '')) if 'Credits' in match_row else ''
                            
                            basket_assignments = pd.concat([
                                basket_assignments,
                                pd.DataFrame([{
                                    'Basket Code': basket_code,
                                    'Course Code': course_code,
                                    'Course Name': course_name,
                                    'Department': dept,
                                    'LTPSC': ltpsc,
                                    'Credits': credits,
                                    'Instructor': instructor
                                }])
                            ], ignore_index=True)
                else:
                    # No 7th sem sheet found - create empty rows for each basket
                    if not baskets.empty:
                        for _, basket_row in baskets.iterrows():
                            basket_code = str(basket_row.get('Course Code', ''))
                            basket_assignments = pd.concat([
                                basket_assignments,
                                pd.DataFrame([{
                                    'Basket Code': basket_code,
                                    'Course Code': '',
                                    'Course Name': '',
                                    'Department': '',
                                    'LTPSC': '',
                                    'Credits': '',
                                    'Instructor': ''
                                }])
                            ], ignore_index=True)
                
                basket_assignments.to_excel(w, sheet_name='Basket_Assignments', index=False)
                print(f"    SUCCESS: Basket assignments sheet created with {len(basket_assignments)} entries")
                
                print(f"\nSUCCESS: Created {filename}")
                print(f"  - Unified timetable with baskets (9:00 AM - 5:30 PM, 2 classes per basket)")
                print(f"  - Basket assignments sheet")
            
            return True
            
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False